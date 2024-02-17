#!coding=utf8
import openpyxl
import xlrd
import time

workbook_cache = {}


class OpenpySheet(object):
	def __init__(self, sheet):
		self.sheet = sheet
		start_time = time.time()
		self.values = tuple(self.sheet.values)
		# print 'load time', time.time() - start_time

	def get_row(self, row_id):
		return self.values[row_id - 1]

	@property
	def max_row(self):
		return len(self.values)

	@property
	def max_column(self):
		return len(self.values[0])
	
	@property
	def header(self):
		return self.values[0] if len(self.values) else []

	def iter_rows(self, min_row=1):
		if min_row == 1:
			return self.values
		else:
			return self.values[min_row - 1:]

	def get_rows(self):
		return self.sheet.get_rows()

	def cell(self, row, column):
		return self.values[row - 1][column - 1]

	@property
	def true_row_count(self):
		row_count = 0
		for row in self.iter_rows():
			for cell in row:
				if cell:
					row_count += 1
				break
		row_count -= 1  # 去掉第一行
		return row_count

	def row_values(self, n_row):
		row = self.values[n_row]
		for col in xrange(len(row)-1, -1, -1):
			if row[col] is None:
				return row[:col]
		return row

	@property
	def nrows(self):
		return len(self.values)


class XlrdSheet(OpenpySheet):
	def __init__(self, sheet):
		self.sheet = sheet
		self.values = self.sheet._cell_values

	def row_values(self, n_row):
		return self.sheet.row_values(n_row)


class OpenpyxlUtil(object):
	"""封装了openpyxl的一些接口，方便使用"""

	def __init__(self, file_name, read_only=True, use_xlrd=True):
		self.file_name = file_name
		self.use_xlrd = use_xlrd
		if (file_name, read_only, use_xlrd) in workbook_cache:
			self.workbook = workbook_cache[(file_name, read_only, use_xlrd)]
		else:
			if use_xlrd:
				self.workbook = xlrd.open_workbook(file_name)
				workbook_cache[(file_name, read_only, use_xlrd)] = self.workbook
			else:
				self.workbook = openpyxl.load_workbook(file_name, read_only=read_only, data_only=True)
				workbook_cache[(file_name, read_only, use_xlrd)] = self.workbook
			self.workbook.sheet_cache = {}

	def get_sheet_by_name(self, sheet_name):
		if not self.workbook:
			return None
		else:
			if sheet_name in self.workbook.sheet_cache:
				return self.workbook.sheet_cache[sheet_name]
			else:
				if self.use_xlrd:
					sheet_data = XlrdSheet(self.workbook.sheet_by_name(sheet_name))
				else:
					sheet_data = OpenpySheet(self.workbook.get_sheet_by_name(sheet_name))
				self.workbook.sheet_cache[sheet_name] = sheet_data
				return sheet_data

	def get_sheet_by_index(self, index):
		if not self.workbook:
			return None
		else:
			sheet_name = self.workbook.sheet_names()[index] if self.use_xlrd else self.workbook.sheetnames[index]
			return self.get_sheet_by_name(sheet_name)

	def get_row(self, sheet_name, row_id):
		"""获取第row_id行的所有数据"""
		return self.get_sheet_by_name(sheet_name).get_row(row_id)

	@staticmethod
	def check_row_data(row_data):
		if not row_data[0]:
			return False
		if row_data[0] == 1:
			return True
		else:
			return False

	def save(self, file_name):
		return self.workbook.save(file_name)


def write_row(sheet, row_index, row_content):
	for n_col, unit in enumerate(row_content):
		sheet.write(row_index, n_col, unit)